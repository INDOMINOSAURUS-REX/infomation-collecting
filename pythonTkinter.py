from tkinter import *
from openpyxl import *

LB = load_workbook("form user inputs.xlsm")
sheet = LB.active

current_row = sheet.max_row
current_column = sheet.max_column
#gender = str

def enter():
    name = ME_name.get()
    ME_name.delete(0, END)
    sheet.cell(row=current_row + 1, column=1).value = name

    age = ME_age.get()
    ME_age.delete(0, END)
    sheet.cell(row=current_row+1,column=2).value = age

    address = ME_address.get()
    ME_address.delete(0, END)
    sheet.cell(row=current_row+1, column=3).value = address

    birth_year=ME_birthyear.get()
    ME_birthyear.delete(0, END)
    sheet.cell(row=current_row+1, column=4).value = birth_year

    mobile_number = ME_Mnumber.get()
    ME_Mnumber.delete(0, END)
    sheet.cell(row=current_row+1, column=5).value = mobile_number
    def checkbox():
        if MC_genderMale.getboolean(True):
            gender = "Male"
            print(gender)
            sheet.cell(row=current_row + 1, column=6).value = gender
        elif MC_genderFeMale.getboolean(True):
            gender = "Female"
            print(gender)
            sheet.cell(row=current_row + 1, column=6).value = gender
        else:
            print("Error")
    LB.save("form user inputs.xlsm")
    checkbox()


# making variables
name = str
age = int
address = str
birth_year = str
mobile_number = str
gender = str
root = Tk()
# root.geometry("400x400")
root.title("Form 0.1")
root.geometry("469x600")
iconphoto = PhotoImage(file="icons8-form-64.png")
root.iconphoto(True, iconphoto)

icon1 = PhotoImage(file="icons8-form-80.png")
ML_icon = Label(root, image=icon1).grid(row=0, column=1, padx=190)


# funshions DEF
def enter1():
    ME_name.get()
    ME_name.delete(0, END)


# MAKING ALL STUFF
My_label1 = Label(root, text="Hello Welcome to our first Form.", font="Vera 15")
ML_name = Label(root, text="What is your name?")
Label(root, text="Enter your name:").grid(row=3, column=0, columnspan=1)
ME_name = Entry(root, )
# MB_nameEnter = Button(root, command=enter1, text="Enter", padx=15, pady=1)

# SPACES BETWEEN WIDGETS
MS_space1 = Label(root).grid(row=6, column=0)
MS_space2 = Label(root).grid(row=10, column=0)
MS_space3 = Label(root).grid(row=13, column=0)
MS_space4 = Label(root).grid(row=16, column=0)
MS_space5 = Label(root).grid(row=19, column=0)

# GRIDING ALL THINGS
My_label1.grid(row=0, column=0, sticky=W, padx=40, pady=20, columnspan=2)
ML_name.grid(row=2, column=0, sticky=W, pady=1, padx=20, columnspan=1)
ME_name.grid(row=3, column=1, sticky=W, ipadx=10, columnspan=1)


# MB_nameEnter.grid(row=5, column=0, )


# AFTER ALL MUST BE LOOPING!

# deffing age button!
def enter2():
    ME_age.get()
    ME_age.delete(0, END)


ML_age = Label(root, text="How old are you?", )
Label(root, text="Enter your age:").grid(row=8, column=0, padx=2)
ME_age = Entry(root, )
# MB_ageEnter = Button(root, command=enter2, text="Enter", padx=15, pady=1)
# gridding asking the age
ML_age.grid(row=7, column=0, sticky=W, padx=20)
ME_age.grid(row=8, column=1, ipadx=10, sticky=W, )
# MB_ageEnter.grid(row=9, column=0, )

# making address input
ML_address = Label(root, text="Enter your home address.")
Label(root, text="Home:").grid(row=12, column=0, )
ME_address = Entry(root, )
# griding address inputs
ML_address.grid(row=11, column=0, )
ME_address.grid(row=12, column=1, ipadx=10, sticky=W)

# making birth year input
ML_birthyear = Label(root, text="Enter your birth year.")
Label(root, text="Eg;1/1/2021 :").grid(row=15, column=0)
ME_birthyear = Entry(root, )
# griding birth year inputs
ML_birthyear.grid(row=14, column=0)
ME_birthyear.grid(row=15, column=1, ipadx=10, sticky=W)

# making mobile number input
ML_Mnumber = Label(root, text="Enter your mobile number.")
Label(root, text="No:").grid(row=18, column=0)
ME_Mnumber = Entry(root, )
# griding mobile number inputs
ML_Mnumber.grid(row=17, column=0)
ME_Mnumber.grid(row=18, column=1, ipadx=10, sticky=W)

# making gender check box
ML_checkBox = Label(root, text="Tick the check box.")
Label(root, text="Gender").grid(row=21, column=0)
MC_genderMale = Checkbutton(root, text="Male")
MC_genderFeMale = Checkbutton(root, text="Female")

# griding gender
ML_checkBox.grid(row=20, column=0)
MC_genderMale.grid(row=21, column=1, sticky=W, )
MC_genderFeMale.grid(row=22, column=1, sticky=W, )
# defining enter button


# last enter button
enter_button = Button(root, command=enter, text="Enter", padx=30, pady=10, border=0.5).grid(row=24, column=1, sticky=S)

root.mainloop()
